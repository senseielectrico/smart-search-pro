"""
Excel exporter with multi-sheet support, formatting, and statistics.
"""

import time
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from .base import BaseExporter, ExportConfig, ExportError, ExportStats
except ImportError:
    from base import BaseExporter, ExportConfig, ExportError, ExportStats


class ExcelExporter(BaseExporter):
    """
    Export search results to Excel format (.xlsx).

    Features:
    - Multiple sheets (by category, extension, or all results)
    - Professional formatting
    - Auto-width columns
    - Freeze panes
    - Excel tables with filters
    - Summary sheet with statistics
    - Charts (optional)
    - Cell hyperlinks for file paths
    """

    def __init__(self, config: Optional[ExportConfig] = None):
        """
        Initialize Excel exporter.

        Args:
            config: Export configuration

        Raises:
            ImportError: If openpyxl is not available
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install with: pip install openpyxl"
            )

        super().__init__(config)

        # Excel-specific options
        self.split_by = self.config.options.get("split_by", None)  # None, "extension", "folder"
        self.include_summary = self.config.options.get("include_summary", True)
        self.freeze_panes = self.config.options.get("freeze_panes", True)
        self.auto_filter = self.config.options.get("auto_filter", True)
        self.use_tables = self.config.options.get("use_tables", True)
        self.add_hyperlinks = self.config.options.get("add_hyperlinks", True)
        self.max_rows_per_sheet = self.config.options.get("max_rows_per_sheet", 1000000)

    def export(self, results: List) -> ExportStats:
        """
        Export results to Excel file.

        Args:
            results: Search results to export

        Returns:
            Export statistics

        Raises:
            ExportError: If export fails
        """
        start_time = time.time()

        # Validate output path
        if not self.config.output_path:
            raise ExportError("Output path is required for Excel export")

        output_path = Path(self.config.output_path)
        self._validate_output_path(output_path)

        # Prepare results
        results = self._prepare_export(results)

        try:
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Split results if requested
            if self.split_by:
                grouped = self._group_results(results)
                self._export_grouped(wb, grouped)
            else:
                self._export_sheet(wb, "Results", results)

            # Add summary sheet
            if self.include_summary:
                self._add_summary_sheet(wb, results)

            # Save workbook
            self._report_progress(90, 100, "Saving workbook")
            wb.save(output_path)

            self.stats.exported_records = len(results)

        except Exception as e:
            self.stats.errors += 1
            raise ExportError(f"Failed to export Excel: {e}") from e

        finally:
            self.stats.duration_seconds = time.time() - start_time

        return self._finalize_stats(output_path)

    def _group_results(self, results: List) -> Dict[str, List]:
        """
        Group results by specified criteria.

        Args:
            results: Results to group

        Returns:
            Dictionary of grouped results
        """
        grouped = defaultdict(list)

        for result in results:
            if self.split_by == "extension":
                key = result.extension.upper() if result.extension else "NO_EXT"
            elif self.split_by == "folder":
                key = Path(result.path).name or "ROOT"
            elif self.split_by == "type":
                key = "Folders" if result.is_folder else "Files"
            else:
                key = "All"

            grouped[key].append(result)

        return dict(grouped)

    def _export_grouped(self, wb: Workbook, grouped: Dict[str, List]) -> None:
        """
        Export grouped results to multiple sheets.

        Args:
            wb: Workbook instance
            grouped: Grouped results
        """
        total_sheets = len(grouped)
        for i, (name, results) in enumerate(sorted(grouped.items())):
            # Sanitize sheet name (Excel limits)
            sheet_name = self._sanitize_sheet_name(name)
            self._export_sheet(wb, sheet_name, results)
            self._report_progress(i + 1, total_sheets, f"Creating sheet: {sheet_name}")

    def _export_sheet(self, wb: Workbook, name: str, results: List) -> None:
        """
        Export results to a single sheet.

        Args:
            wb: Workbook instance
            name: Sheet name
            results: Results for this sheet
        """
        ws = wb.create_sheet(title=name)

        # Write headers
        for col_idx, col_name in enumerate(self.config.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            self._style_header(cell)

        # Write data
        for row_idx, result in enumerate(results[:self.max_rows_per_sheet], start=2):
            for col_idx, col_name in enumerate(self.config.columns, start=1):
                value = self._format_value(result, col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Add hyperlinks for file paths
                if self.add_hyperlinks and col_name == "full_path" and value:
                    try:
                        cell.hyperlink = f"file:///{value}"
                        cell.style = "Hyperlink"
                    except Exception:
                        pass  # Skip if path is invalid

        # Apply formatting
        self._format_sheet(ws, len(results))

    def _add_summary_sheet(self, wb: Workbook, results: List) -> None:
        """
        Add summary statistics sheet.

        Args:
            wb: Workbook instance
            results: All results
        """
        ws = wb.create_sheet(title="Summary", index=0)

        # Calculate statistics
        total_files = sum(1 for r in results if not r.is_folder)
        total_folders = sum(1 for r in results if r.is_folder)
        total_size = sum(r.size for r in results if not r.is_folder)
        extensions = defaultdict(int)
        for r in results:
            if r.extension:
                extensions[r.extension.upper()] += 1

        # Write summary
        row = 1
        ws.cell(row, 1, "Search Results Summary").font = Font(size=16, bold=True)
        row += 2

        # General stats
        stats = [
            ("Total Results", len(results)),
            ("Files", total_files),
            ("Folders", total_folders),
            ("Total Size", self._format_size_human(total_size)),
            ("Unique Extensions", len(extensions)),
        ]

        for label, value in stats:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1

        # Top extensions
        if extensions:
            row += 2
            ws.cell(row, 1, "Top File Types").font = Font(size=12, bold=True)
            row += 1

            ws.cell(row, 1, "Extension").font = Font(bold=True)
            ws.cell(row, 2, "Count").font = Font(bold=True)
            row += 1

            for ext, count in sorted(extensions.items(), key=lambda x: x[1], reverse=True)[:10]:
                ws.cell(row, 1, ext)
                ws.cell(row, 2, count)
                row += 1

        # Auto-width columns
        for col in [1, 2]:
            ws.column_dimensions[get_column_letter(col)].width = 20

    def _format_sheet(self, ws, num_rows: int) -> None:
        """
        Apply formatting to sheet.

        Args:
            ws: Worksheet instance
            num_rows: Number of data rows
        """
        # Auto-width columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width

        # Freeze panes
        if self.freeze_panes and num_rows > 0:
            ws.freeze_panes = "A2"

        # Auto filter
        if self.auto_filter and num_rows > 0:
            ws.auto_filter.ref = ws.dimensions

        # Table formatting
        if self.use_tables and num_rows > 0:
            try:
                table_ref = f"A1:{get_column_letter(len(self.config.columns))}{num_rows + 1}"
                table = Table(displayName=f"Table_{ws.title.replace(' ', '_')}", ref=table_ref)
                style = TableStyleInfo(
                    name="TableStyleMedium2",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                table.tableStyleInfo = style
                ws.add_table(table)
            except Exception:
                pass  # Skip if table creation fails

    def _style_header(self, cell) -> None:
        """
        Apply header styling.

        Args:
            cell: Cell to style
        """
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            bottom=Side(style="thin", color="000000")
        )

    @staticmethod
    def _sanitize_sheet_name(name: str) -> str:
        """
        Sanitize sheet name for Excel compatibility.

        Args:
            name: Original name

        Returns:
            Sanitized name
        """
        # Excel sheet name restrictions
        invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
        for char in invalid_chars:
            name = name.replace(char, '_')

        # Max 31 characters
        return name[:31]

    @staticmethod
    def _format_size_human(size: int) -> str:
        """Format size in human-readable format."""
        for unit in ["B", "KB", "MB", "GB", "TB"]:
            if size < 1024.0:
                return f"{size:.2f} {unit}"
            size /= 1024.0
        return f"{size:.2f} PB"
